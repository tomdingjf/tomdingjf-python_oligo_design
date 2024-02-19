# Source Generated with Decompyle++
# File: oligo_design.pyc (Python 3.6)

import urllib.request
import urllib.parse
import openpyxl
import os
import shutil
import sys

ORIGINAL_FILE = '批量引物设计.xlsx'
ORIGINAL_SHEET1 = '订单录入'
ORIGINAL_SHEET2 = '载体对应接头序列'
RESULT_FILE = '订单处理结果汇总.xlsx'
RESULT_SHEET = '订单处理结果汇总'
ERRLOG_FILE = '处理失败订单统计.xlsx'
ERRLOG_SHEET = '失败订单信息'
ERR_REASON1 = '找不到载体对应的接头'
ERR_REASON2 = '网页处理失败'


def url_open(job_name, sequence, left_primer, right_primer, tm_max, ol_min):
    url = 'http://192.168.4.251/cgi-bin/oligo_design/oligo_design.cgi'  # 定义url
    data = {
        'job_name': job_name,  # 作业名称
        'left_primer': left_primer,  # 左引物
        'ol_max': '59',  # 最大退火长度
        'ol_min': ol_min,  # 最小退火长度
        'op': 'Send',  # 操作
        'ov_min': '17',  # 最小熔点
        'right_primer': right_primer,  # 右引物
        'sequence': sequence,  # 序列
        'tm_max': tm_max,  # 最大退火温度
        'tm_min': '30'  # 最小退火温度
    }
    download_addrs = []  # 下载地址列表
    data = urllib.parse.urlencode(data).encode('utf-8')  # 编码为utf-8格式
    response = urllib.request.urlopen(url, data)  # 发送请求
    html = response.read().decode('utf-8')  # 读取响应内容
    b = 0
    for index in range(2):  # 循环两次
        a = html.find('<a href="', b)  # 查找<a href="的索引
        if -1 == a:  # 如果没有找到
            download_addrs.append(ERR_REASON2)  # 添加错误原因
            break
        b = html.find('"', a + 9)  # 查找双引号的索引
        download_addrs.append(html[a + 9:b])  # 添加下载地址
    print("调用1")
    return download_addrs  # 返回下载地址列表


def save_files(job_name, download_addrs, order):
    print("调用2")
    os.mkdir(job_name)


# WARNING: Decompyle incomplete


def parse_vec2jo(wb):
    sheet = wb[ORIGINAL_SHEET2]
    vec2jo = {}
    for index in range(sheet.max_row - 1):
        vec2jo[sheet['A%d' % (index + 2)].value] = [
            sheet['B%d' % (index + 2)].value,
            sheet['C%d' % (index + 2)].value]
    print("调用3")
    return vec2jo


def parse_orders(wb, vec2jo):
    sheet = wb[ORIGINAL_SHEET1]
    orders = {}
    column_name = {}
    print("调用4头")
    for each in sheet['a']:
        print(f"each.value:{each.value}")
        print(f"each.value and:{each.value and '序列'}")
        print("调用4头1")

        if not '订单号' == each.value and '序列' == each.value and '载体' == each.value and '载体酶切位点' == each.value and "5'接头" == each.value and '3’接头' == each.value and '订单类型' == each.value and '亚克隆5’附加' == each.value:
            print("调用4头2")
            if '亚克隆3’附加' == each.value:
                column_name[each.value] = each.column
            for index in range(sheet.max_row - 1):
                job_name = sheet['%c%d' % (column_name['订单号'], index + 2)].value
                print(f"job_name:{job_name}")
                sequence = sheet['%c%d' % (column_name['序列'], index + 2)].value
                vector = sheet['%c%d' % (column_name['载体'], index + 2)].value
                site = sheet['%c%d' % (column_name['载体酶切位点'], index + 2)].value
                left_primer = sheet['%c%d' % (column_name["5'接头"], index + 2)].value
                right_primer = sheet['%c%d' % (column_name['3’接头'], index + 2)].value
                job_type = sheet['%c%d' % (column_name['订单类型'], index + 2)].value
                left_subadd = sheet['%c%d' % (column_name['亚克隆5’附加'], index + 2)].value
                right_subadd = sheet['%c%d' % (column_name['亚克隆3’附加'], index + 2)].value
                if not job_name:
                    continue
                if not left_primer or not right_primer:
                    if vector in vec2jo:
                        left_primer = vec2jo[vector][0]
                        right_primer = vec2jo[vector][1]
                    else:
                        left_primer = ERR_REASON1
                        right_primer = ERR_REASON1
                if not left_subadd:
                    left_subadd = 20
                if not right_subadd:
                    right_subadd = 20
                sequence = sequence.upper()
                left_primer = left_primer.upper()
                right_primer = right_primer.upper()
                orders[job_name] = [
                    sequence,
                    left_primer,
                    right_primer,
                    vector,
                    site,
                    job_type,
                    left_subadd,
                    right_subadd]
    print("调用4头尾")
    print(f"orders:{orders}")
    return orders


def save_errlog(err1, err2):
    """
    保存错误日志
    :param err1: 第一种错误类型及其相关信息的字典
    :param err2: 第二种错误类型及其相关信息的字典
    """
    wb = openpyxl.Workbook()  # 创建一个工作簿
    sheet = wb.active  # 获取活动工作表
    sheet.title = ERRLOG_SHEET  # 设置工作表标题
    sheet['A1'] = '订单号'  # 设置表头
    sheet['B1'] = '序列'
    sheet['C1'] = "5'接头"
    sheet['D1'] = "3'接头"
    sheet['E1'] = '失败原因'

    count = 2  # 计数器，用于填写数据行的列索引
    for each in err1:  # 遍历err1字典的键值对
        sheet['A%d' % count].value = each  # 填充订单号
        sheet['B%d' % count].value = err1[each][0]  # 填充序列
        sheet['C%d' % count].value = err1[each][1]  # 填充5'接头
        sheet['D%d' % count].value = err1[each][2]  # 填充3'接头
        sheet['E%d' % count].value = ERR_REASON1  # 填充失败原因
        count += 1  # 计数器自增

    for each in err2:  # 遍历err2字典的键值对
        sheet['A%d' % count].value = each  # 填充订单号
        sheet['B%d' % count].value = err2[each][0]  # 填充序列
        sheet['C%d' % count].value = err2[each][1]  # 填充5'接头
        sheet['D%d' % count].value = err2[each][2]  # 填充3'接头
        sheet['E%d' % count].value = ERR_REASON2  # 填充失败原因
        count += 1  # 计数器自增

    if 2 != count:  # 判断数据行数是否为2，即是否填写完整
        wb.save(ERRLOG_FILE)  # 保存工作簿到文件
    print("调用5")


def file_del():
    """
    删除指定目录下的所有文件和子目录

    Args:

    Returns:

    """
    # 获取指定目录下的所有文件和子目录列表
    file_list = os.listdir()

    # 移除原始文件
    file_list.remove(ORIGINAL_FILE)

    # 移除当前脚本文件名
    file_list.remove(os.path.basename(sys.argv[0]))

    # 遍历文件和子目录列表
    for each in file_list:
        # 如果是子目录，则删除该子目录及其内容
        if os.path.isdir(each):
            shutil.rmtree(each)
        else:
            # 如果是文件，则删除该文件
            os.remove(each)
    print("调用6")


def file_proc(file_str):
    # 定义一个空字典，用于存储文件信息
    file_dict = {}
    # 统计换行符的数量，用于split后得到的行数
    count = file_str.count('\n')
    # 根据换行符数量split文件内容，并去除首行
    file_list = file_str.split('\n', count - 1)

    # 遍历文件内容列表
    for each in file_list:
        # 如果当前行与首行相同，则跳过
        if each == file_list[0]:
            continue
        # 将当前行按制表符分割成4部分，最多分割3次
        temp = each.split('\t', 3)
        # 将第一部分作为键，将第二、第三、第四部分作为值存储到字典中
        file_dict[temp[0]] = [
            temp[1],
            temp[2],
            temp[3]
        ]
    print("调用7")
    # 返回存储文件信息的字典
    return file_dict


def files_all_proc(file_str_all):
    """
    对给定的文件内容进行处理并写入到Excel表格中

    参数:
    file_str_all (list): 包含所有文件内容的列表

    返回:
    无
    """

    # 创建一个新的Excel工作簿
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = RESULT_SHEET

    # 在第一行写入表头信息
    sheet['A1'] = 'Name'
    sheet['B1'] = "Sequence(5'->3)"
    sheet['C1'] = 'Length'
    sheet['D1'] = 'Position'

    # 从第二行开始逐行写入文件处理结果
    count = 2
    for each in file_str_all:
        # 对每个文件内容进行处理，得到文件字典
        file_dict = file_proc(each)
        for key in file_dict:
            # 将处理结果写入表格中
            sheet['A%d' % count].value = key
            sheet['B%d' % count].value = file_dict[key][0]
            sheet['C%d' % count].value = file_dict[key][1]
            sheet['D%d' % count].value = file_dict[key][2]
            count += 1

    # 保存Excel工作簿到指定文件
    print("调用8")
    wb.save(RESULT_FILE)


def save_files_subclone(job_name, order):
    # 获取序列、左引物、右引物、载体、位点、左子添加、右子添加信息
    sequence = order[0]
    left_primer = order[1]
    right_primer = order[2]
    vector = order[3]
    site = order[4]
    left_subadd = order[6]
    right_subadd = order[7]

    # 在序列中找到左子添加位点前的第一次出现的G和C的位置
    index_g = sequence.find('G', left_subadd - 1)
    index_c = sequence.find('C', left_subadd - 1)

    # 获取左边界位置
    index = min(index_g, index_c)

    # 构建前导序列
    pf_str = sequence[:index + 1]
    pf_str = left_primer + pf_str

    # 在序列中找到右子添加位点后最后一次出现的G和C的位置
    index_g = sequence.rfind('G', 0, -(right_subadd - 1))
    index_c = sequence.rfind('C', 0, -(right_subadd - 1))

    # 获取右边界位置
    index = max(index_g, index_c)

    # 构建后随序列
    pr_str = sequence[index:]
    pr_str = pr_str + right_primer

    # 定义要进行翻译的字符为'ACGT'
    in_str = 'ACGT'
    # 定义字符映射表，将'ACGT'翻译为'TGCA'
    out_str = 'TGCA'

    # 将后随序列进行翻译，并将字符进行映射
    pr_str = pr_str.translate(str.maketrans(in_str, out_str))

    # 将后随序列进行反转
    pr_str = pr_str[::-1]

    # 构建输出的行1、行2和行3
    line1 = "Name\tSequence(5'->3')\t\t\t\n"
    line2 = job_name + ' PF\t' + pf_str + '\t\t\t\n'
    line3 = job_name + ' PR\t' + pr_str + '\t\t\t\n'

    # 拼接行1、行2和行3为最终的输出行
    line = line1 + line2 + line3

    # 创建以作业名称为名称的目录
    os.mkdir(job_name)
    print("调用9")


# WARNING: Decompyle incomplete


def oligo_design():
    # 加载原始工作簿
    wb = openpyxl.load_workbook(ORIGINAL_FILE)

    # 解析vec2jo
    vec2jo = parse_vec2jo(wb)
    print(f"vec2jo:{vec2jo}")
    # 解析orders
    orders = parse_orders(wb, vec2jo)
    print(f"orders:{orders}")
    # 初始化错误记录字典
    err1 = {}
    err2 = {}

    # 初始化文件字符串列表
    file_str_all = []

    # 定义tm_max和ol_min列表
    tm_max = [
        '68',
        '80',
        '80']

    ol_min = [
        '55',
        '55',
        '30']

    # 删除文件
    file_del()
    print(99)
    # 遍历orders字典
    for each in orders:
        # 如果ERR_REASON1等于orders[each][1]或orders[each][2]，则将orders[each]添加到err1字典中，然后继续下一次循环
        if ERR_REASON1 == orders[each][1] or ERR_REASON1 == orders[each][2]:
            err1[each] = orders[each]
            continue

        # 获取job_name、sequence、left_primer、right_primer、job_type和subadd
        job_name = each[3:]
        print(f"job_name:{job_name}")
        sequence = orders[each][0]
        left_primer = orders[each][1]
        right_primer = orders[each][2]
        job_type = orders[each][5]
        subadd = orders[each][6]

        # 如果job_type是'亚克隆'或'载体构建'，则调用save_files_subclone函数保存文件，并将返回的文件字符串添加到file_str_all列表中，然后继续下一次循环
        if '亚克隆' == job_type or '载体构建' == job_type:
            file_str = save_files_subclone(job_name, orders[each])
            file_str_all.append(file_str)
            continue

        # 遍历tm_max列表
        for index in range(len(tm_max)):
            # 调用url_open函数下载文件，并将返回的下载地址添加到download_addrs列表中
            print(job_name, sequence, left_primer, right_primer, tm_max[index], ol_min[index])
            download_addrs = url_open(job_name, sequence, left_primer, right_primer, tm_max[index], ol_min[index])

            # 如果ERR_REASON2不等于download_addrs[0]，则调用save_files函数保存文件，并将返回的文件字符串添加到file_str_all列表中，然后跳出循环
            if ERR_REASON2 != download_addrs[0]:
                file_str = save_files(job_name, download_addrs, orders[each])
                file_str_all.append(file_str)
                break

            # 将orders[each]添加到err2字典中
            err2[each] = orders[each]

    # 调用files_all_proc函数处理file_str_all列表中的文件
    files_all_proc(file_str_all)

    # 调用save_errlog函数保存错误记录err1和err2到错误日志文件中
    save_errlog(err1, err2)
    print("调用10")


if __name__ == '__main__':
    oligo_design()
